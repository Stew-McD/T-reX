"""
MakeCustomDatabase Module
=========================

This module contains functions for creating an xlsx representation of a Brightway2 database 
and importing it into Brightway2. 

Main functions:
- dbWriteExcel: Creates an xlsx file representing a custom Brightway2 database.
- dbExcel2BW: Imports the custom database (created by dbWriteExcel) into Brightway2.

"""

import glob
import os

import bw2data as bd
import bw2io as bi
from openpyxl import Workbook, load_workbook
from config.user_settings import (
    db_wmf_name,
    dir_databases_wasteandmaterial,
    dir_searchmaterial_results,
    dir_searchwaste_results,
    project_wmf,
)


def get_files_from_tree(dir_searchmaterial_results, dir_searchwaste_results):
    """
    Collects filenames from the SearchMaterial and SearchWasteResults directories.

    :param dir_searchmaterial_results: Directory path for SearchMaterial results.
    :param dir_searchwaste_results: Directory path for SearchWasteResults.
    :return: Sorted list of filenames.
    """
    # Get files in the SearchMaterial/*/grouped/ directory
    search_material_files = glob.glob(
        os.path.join(dir_searchmaterial_results, "*", "grouped", "*")
    )

    # Get files in the SearchWasteResults/* directory
    search_waste_files = glob.glob(os.path.join(dir_searchwaste_results, "*", "*"))

    all_files = search_material_files + search_waste_files

    # Extract only the filename without the suffix
    names = sorted(
        set(os.path.splitext(os.path.basename(file))[0] for file in all_files)
    )

    return names


def dbWriteExcel():
    """
    Create an xlsx file representing a custom Brightway2 database.

    This function generates an Excel file which represents a custom database for Brightway2,
    using predefined directory and database settings.

    :return: Path to the generated xlsx file.
    """

    if not os.path.isdir(dir_databases_wasteandmaterial):
        os.makedirs(dir_databases_wasteandmaterial)

    xl_filename = dir_databases_wasteandmaterial / f"{db_wmf_name}.xlsx"

    # delete existing file if it exists
    if os.path.isfile(xl_filename):
        os.remove(xl_filename)

    # create new file and write header
    print(f"\n\n*** Writing custom database file: {db_wmf_name}\n")

    xl = Workbook()
    xl_db = xl.active
    xl_db["A1"] = "Database"
    xl_db["B1"] = db_wmf_name
    xl_db["A2"] = ""

    xl.save(xl_filename)

    # open existing file and append to it
    print(f"\n\n*** Appending to existing custom database file: {db_wmf_name}\n")

    xl = load_workbook(xl_filename)
    xl_db = xl.active

    count = 0
    names = get_files_from_tree(dir_searchmaterial_results, dir_searchwaste_results)
    for NAME in names:
        count += 1
        CODE = NAME
        UNIT = determine_unit_from_name(NAME)

        if "Waste" in NAME:
            TYPE = "waste"
            CODE = (
                NAME.replace("WasteFootprint_", "")
                .capitalize()
                .replace("kilogram", "(kg)")
                .replace("cubicmeter", "(m3)")
                .replace("-", " ")
            )
        elif "Material" in NAME:
            TYPE = "natural resource"
            CODE = NAME.replace("MaterialFootprint_", "").capitalize()
        else:
            TYPE = "?"

        db_entry = {
            "Activity": NAME,
            "categories": "water, air, land",
            "code": CODE,
            "unit": UNIT,
            "type": TYPE,
        }

        print("\t Appending:", NAME)
        for key, value in db_entry.items():
            row = [key, str(value)]
            xl_db.append(row)
        xl_db.append([""])

    xl.save(xl_filename)
    print(
        f"\n ** Added {count} entries to the xlsx for the custom waste and material database:\n\t{db_wmf_name}"
    )

    return


def determine_unit_from_name(name):
    """
    Determine the unit based on the name.

    :param name: The name from which to infer the unit.
    :return: The inferred unit as a string.
    """
    if "kilogram" in name:
        return "kilogram"
    elif "cubicmeter" in name or "water" in name or "gas" in name:
        return "cubic meter"
    elif "electricity" in name:
        return "kilowatt hour"
    elif "Material" in name:
        return "kilogram"
    else:
        return ""


def dbExcel2BW():
    """
    Import the custom database (created by dbWriteExcel) into Brightway2.

    This function imports a custom Brightway2 database from an Excel file into the Brightway2 software,
    making it available for further environmental impact analysis.

    :return: None
    """
    print(
        f"\n** Importing the custom database {db_wmf_name}**\n\t to the brightway2 project: {project_wmf}"
    )

    xl_filename = dir_databases_wasteandmaterial / f"{db_wmf_name}.xlsx"
    bd.projects.set_current(project_wmf)

    if db_wmf_name not in bd.databases:
        # imports the custom database into BW2
        print("\n** Running BW2io ExcelImporter **\n")
        imp = bi.ExcelImporter(xl_filename)
        bi.create_core_migrations()  # needed to stop it from occasional crashing
        imp.apply_strategies()
        imp.statistics()
        imp.write_database()

        db_wmf = bd.Database(db_wmf_name)
        db_wmf.register()

        db_dict = db_wmf.metadata
        print("\n** Database metadata **")
        for key, value in db_dict.items():
            print(f"{key}: {value}")

    else:
        print(f"\n** Database {db_wmf_name} already exists **\n")
        db_wmf = bd.Database(db_wmf_name)
        imp = bi.ExcelImporter(xl_filename)

        for act in imp:
            if act["name"] not in db_wmf:
                print(
                    f"\t|-  Adding activity: {act['name']:<40} ----> \t '{db_wmf_name}'  -|"
                )
                db_wmf.new_activity(act)
            else:
                print(f"\t {act['name']} already exists in {db_wmf_name}")

    print("\n*** Great success! ***")

    return None


if __name__ == "__main__":
    dbWriteExcel()
    dbExcel2BW()